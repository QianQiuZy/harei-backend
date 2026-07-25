from datetime import date
from io import BytesIO
from typing import ClassVar

from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ConfigDict
import pytest

from app.schemas.auth import LoginResponse


class ResponseModel(BaseModel):
    model_config: ClassVar[ConfigDict] = ConfigDict(frozen=True)


class CreatedSongResponse(ResponseModel):
    song_id: int
    source_key: str


class PublicStreamResponse(ResponseModel):
    id: str | None


class PublicPerformanceResponse(ResponseModel):
    id: str
    stream: PublicStreamResponse


class PublicSongResponse(ResponseModel):
    performances: tuple[PublicPerformanceResponse, ...]


class PublicDetailResponse(ResponseModel):
    item: PublicSongResponse


async def music_headers(client: AsyncClient) -> dict[str, str]:
    response = await client.post(
        "/music-manage/login",
        json={"username": "music", "password": "music-password"},
    )
    session = LoginResponse.model_validate(response.json())
    return {"Authorization": f"Bearer {session.token}"}


def workbook_bytes(rows: list[tuple[str, date | str, str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Workbook has no active worksheet")
    sheet.title = "导入数据"
    sheet.append(["歌名", "日期", "直播标题", "歌切链接"])
    for row in rows:
        sheet.append(row)
    workbook.create_sheet("歌曲列表")
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
async def test_template_contains_named_sheets_and_current_songs(
    client: AsyncClient,
    song_payload: dict[str, str | list[str]],
) -> None:
    headers = await music_headers(client)
    created = await client.post("/music-manage/songs", headers=headers, json=song_payload)
    assert created.status_code == 201

    response = await client.get("/music-manage/performances/template", headers=headers)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["导入数据", "歌曲列表"]
    assert tuple(workbook["导入数据"].values)[0] == ("歌名", "日期", "直播标题", "歌切链接")
    assert tuple(workbook["歌曲列表"].values)[1][0] == song_payload["title"]


@pytest.mark.asyncio
async def test_import_creates_performance_and_parses_bv_id(
    client: AsyncClient,
    song_payload: dict[str, str | list[str]],
) -> None:
    headers = await music_headers(client)
    created = await client.post("/music-manage/songs", headers=headers, json=song_payload)
    created_song = CreatedSongResponse.model_validate(created.json())
    upload = workbook_bytes(
        [
            (
                "Test Song",
                date(2026, 7, 26),
                "测试直播",
                "https://www.bilibili.com/video/BV1ABCDEF123/?p=2",
            )
        ]
    )

    response = await client.post(
        "/music-manage/performances/import",
        headers=headers,
        files={"file": ("performances.xlsx", upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 201
    assert response.json()["imported_count"] == 1
    assert response.json()["affected_song_count"] == 1
    detail = await client.get(f"/music/{created_song.source_key}")
    performance = PublicDetailResponse.model_validate(detail.json()).item.performances[0]
    assert performance.id.startswith("performance_")
    assert performance.stream.id == "BV1ABCDEF123"
    managed = await client.get(f"/music-manage/songs/{created_song.song_id}", headers=headers)
    assert managed.json()["item"]["version"] == 2


@pytest.mark.asyncio
async def test_import_rejects_unknown_song_without_partial_writes(
    client: AsyncClient,
    song_payload: dict[str, str | list[str]],
) -> None:
    headers = await music_headers(client)
    created = await client.post("/music-manage/songs", headers=headers, json=song_payload)
    created_song = CreatedSongResponse.model_validate(created.json())
    upload = workbook_bytes(
        [
            ("Test Song", "2026-07-26", "有效记录", "https://www.bilibili.com/video/BV1ABCDEF123"),
            ("不存在的歌曲", "2026-07-27", "无效记录", "https://www.bilibili.com/video/BV1ZZZZZZZZZ"),
        ]
    )

    response = await client.post(
        "/music-manage/performances/import",
        headers=headers,
        files={"file": ("performances.xlsx", upload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 422
    assert response.json()["detail"]["errors"][0]["row"] == 3
    detail = await client.get(f"/music/{created_song.source_key}")
    assert PublicDetailResponse.model_validate(detail.json()).item.performances == ()


@pytest.mark.asyncio
async def test_import_rejects_corrupt_xlsx(client: AsyncClient) -> None:
    headers = await music_headers(client)

    response = await client.post(
        "/music-manage/performances/import",
        headers=headers,
        files={
            "file": (
                "corrupt.xlsx",
                b"not an xlsx archive",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"]["errors"][0]["code"] == "INVALID_XLSX"
