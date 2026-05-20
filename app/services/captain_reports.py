from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.models.captain import Captain


def build_captains_workbook(rows: Sequence[Captain], month: str) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = f"{month}舰长"

    headers = ["UID", "用户名", "舰长等级", "上舰数量", "上舰时间", "是否红包"]
    worksheet.append(headers)

    column_widths = [len(header) for header in headers]
    for row in rows:
        joined_at = row.joined_at.strftime("%Y-%m-%d %H:%M:%S")
        values = [
            row.user_uid,
            row.username or "",
            row.level,
            row.ship_count,
            joined_at,
            "是" if row.is_red_packet else "否",
        ]
        worksheet.append(values)
        for index, value in enumerate(values):
            value_length = len(str(value))
            if value_length > column_widths[index]:
                column_widths[index] = value_length

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width + 2

    return workbook


def save_captains_xlsx(rows: Sequence[Captain], month: str, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / f"captains_{month}.xlsx"
    workbook = build_captains_workbook(rows, month)
    workbook.save(file_path)
    return file_path
