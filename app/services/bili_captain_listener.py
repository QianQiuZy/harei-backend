from __future__ import annotations

import asyncio
import calendar
import datetime
import http.cookies
import logging
import random
import smtplib
import threading
from dataclasses import dataclass
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any, Dict, Optional

import aiohttp
import blivedm
from aiohttp import ContentTypeError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.core.config import get_settings
from app.db.session import async_session_factory
from sqlalchemy import select
from sqlalchemy.dialects.mysql import insert

from app.models.captain import Captain
from app.models.gift_ranking import GiftRanking

logger = logging.getLogger(__name__)


aiohttp_session: Optional[aiohttp.ClientSession] = None

ROOM_IDS: list[int] = [1820703922]
ROOM_ANCHORS = {1820703922: "花礼Harei"}
ROOM_CLIENTS: Dict[int, blivedm.BLiveClient] = {}
LAST_RECONNECT: Dict[int, datetime.datetime] = {}

ROOM_UIDS: Dict[int, int] = {}

LAST_STATUS: Dict[int, int] = {}
LIVE_INFO: Dict[int, Dict[str, str]] = {}

COOKIE_ALERT_SENT = False


def _send_cookie_invalid_email_async(log_line: str) -> None:
    global COOKIE_ALERT_SENT
    if COOKIE_ALERT_SENT:
        return
    COOKIE_ALERT_SENT = True

    settings = get_settings()
    if not (settings.smtp_host and settings.email_from and settings.email_to):
        logger.warning("[SMTP] 未配置 SMTP_HOST/EMAIL_FROM/EMAIL_TO，跳过 Cookies 告警邮件")
        return

    def _worker() -> None:
        try:
            subject = "B站直播监听 Cookies 失效告警"
            body = (
                "检测到 B 站直播消息 uid=0，疑似 SESSDATA Cookies 已失效。\n\n"
                f"原始日志：{log_line}\n"
                "请尽快检查并更新 BILI_* Cookies 环境变量。"
            )
            msg = MIMEText(body, "plain", "utf-8")
            msg["Subject"] = subject
            msg["From"] = settings.email_from
            msg["To"] = settings.email_to

            if int(settings.smtp_port) == 465:
                with smtplib.SMTP_SSL(settings.smtp_host, settings.smtp_port) as server:
                    if settings.smtp_user and settings.smtp_pass:
                        server.login(settings.smtp_user, settings.smtp_pass)
                    server.sendmail(settings.email_from, [settings.email_to], msg.as_string())
            else:
                with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
                    server.starttls()
                    if settings.smtp_user and settings.smtp_pass:
                        server.login(settings.smtp_user, settings.smtp_pass)
                    server.sendmail(settings.email_from, [settings.email_to], msg.as_string())

            logger.info("[SMTP] Cookies 失效告警邮件已发送")
        except Exception as e:
            logger.error("[SMTP] 发送 Cookies 失效告警失败: %s", e)

    threading.Thread(target=_worker, daemon=True).start()


def _now() -> datetime.datetime:
    return datetime.datetime.now()


def month_str(dt: Optional[datetime.datetime] = None) -> str:
    dt = dt or _now()
    return dt.strftime("%Y%m")


def _next_month_end_minute(now: datetime.datetime) -> datetime.datetime:
    last_day = calendar.monthrange(now.year, now.month)[1]
    target = now.replace(day=last_day, hour=23, minute=59, second=0, microsecond=0)
    if now < target:
        return target
    year = now.year + (1 if now.month == 12 else 0)
    month = 1 if now.month == 12 else now.month + 1
    last_day = calendar.monthrange(year, month)[1]
    return datetime.datetime(year=year, month=month, day=last_day, hour=23, minute=59, second=0)


def _parse_room_ids_from_env(raw: str) -> list[int]:
    out: list[int] = []
    for part in raw.split(","):
        s = part.strip()
        if not s:
            continue
        try:
            out.append(int(s))
        except ValueError:
            continue
    return out


def init_session() -> None:
    settings = get_settings()

    cookies_base = {
        "SESSDATA": settings.bili_sessdata,
        "bili_jct": settings.bili_bili_jct,
        "DedeUserID": settings.bili_dedeuserid,
        "DedeUserID__ckMd5": settings.bili_dedeuserid_ckmd5,
        "sid": settings.bili_sid,
        "buvid3": settings.bili_buvid3,
        "deviceFingerprint": settings.bili_device_fingerprint,
    }

    cookies = http.cookies.SimpleCookie()
    for k, v in cookies_base.items():
        if not v:
            continue
        cookies[k] = v
        cookies[k]["domain"] = "bilibili.com"

    global aiohttp_session
    connector = aiohttp.TCPConnector(ssl=False)
    aiohttp_session = aiohttp.ClientSession(connector=connector)
    aiohttp_session.cookie_jar.update_cookies(cookies)

    logger.info("[session] 已初始化 Cookies：%s", ",".join(cookies.keys()) or "(empty)")

    raw_room_ids = settings.bili_room_ids or ""
    if raw_room_ids.strip():
        ids = _parse_room_ids_from_env(raw_room_ids)
        if ids:
            global ROOM_IDS
            ROOM_IDS = ids
            logger.info("[session] 使用 env 覆盖 ROOM_IDS，总数=%d", len(ROOM_IDS))

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/123.0.0.0 Safari/537.36"
)

LIVE_STATUS_API = "https://api.live.bilibili.com/room/v1/Room/get_status_info_by_uids"
ROOM_INFO_API = "https://api.live.bilibili.com/room/v1/Room/get_info"
GUARD_LIST_API = "https://api.live.bilibili.com/xlive/app-room/v2/guardTab/topListNew"
GUARD_REPORT_RECEIVER = "harei0301@163.com"
GUARD_REPORT_ROOT = Path("download_files")
GUARD_REPORT_RUID = 1048135385
GUARD_REPORT_ROOM_ID = 1820703922


async def _fetch_room_uid(room_id: int) -> Optional[int]:
    if aiohttp_session is None:
        logger.error("[RoomInfo] aiohttp_session 未初始化")
        return None

    url = f"{ROOM_INFO_API}?room_id={room_id}"
    try:
        async with aiohttp_session.get(
            url,
            timeout=5,
            headers={"User-Agent": USER_AGENT, "Referer": "https://live.bilibili.com"},
        ) as resp:
            if resp.status != 200:
                logger.warning("[RoomInfo] room_id=%s get_info HTTP %s", room_id, resp.status)
                return None
            try:
                payload = await resp.json(content_type=None)
            except ContentTypeError:
                text = (await resp.text())[:200]
                logger.warning("[RoomInfo] room_id=%s get_info 非 JSON，前 200 字：%s", room_id, text)
                return None
    except Exception as e:
        logger.error("[RoomInfo] room_id=%s 请求异常: %s", room_id, e)
        return None

    data = payload.get("data") or {}
    uid_raw = data.get("uid")
    try:
        uid = int(uid_raw)
    except (TypeError, ValueError):
        uid = 0

    if uid:
        return uid
    logger.warning("[RoomInfo] room_id=%s uid 获取失败，原始值=%r", room_id, uid_raw)
    return None


async def init_uids_once(max_rounds: int = 5) -> None:
    logger.info("[init] 开始初始化 UID（get_info）")
    for round_idx in range(1, max_rounds + 1):
        missing = [rid for rid in ROOM_IDS if rid not in ROOM_UIDS]
        if not missing:
            logger.info("[init] 所有 UID 已成功获取")
            return
        logger.info("[init] 第 %d/%d 轮获取 UID，待获取房间数=%d", round_idx, max_rounds, len(missing))
        for room_id in missing:
            uid = await _fetch_room_uid(room_id)
            if uid:
                ROOM_UIDS[room_id] = uid
            await asyncio.sleep(0.3)

    missing = [rid for rid in ROOM_IDS if rid not in ROOM_UIDS]
    if missing:
        logger.error("[init] 经过 %d 轮仍有 UID 获取失败，将在状态轮询中跳过: %s", max_rounds, missing)


async def monitor_all_rooms_status() -> None:
    while not ROOM_UIDS:
        logger.info("[LiveStatus] 等待 UID 初始化...")
        await asyncio.sleep(1)

    logger.info("[LiveStatus] UID 初始化完成，可用房间数=%d，启动状态轮询", len(ROOM_UIDS))

    for rid in ROOM_IDS:
        LAST_STATUS.setdefault(rid, 0)
        LIVE_INFO.setdefault(rid, {"live_time": "0000-00-00 00:00:00", "title": ""})

    while True:
        try:
            if aiohttp_session is None:
                logger.error("[LiveStatus] aiohttp_session 未初始化")
                await asyncio.sleep(3)
                continue

            params = [("uids[]", str(uid)) for uid in ROOM_UIDS.values()]
            async with aiohttp_session.get(
                LIVE_STATUS_API,
                params=params,
                timeout=10,
                headers={"User-Agent": USER_AGENT, "Referer": "https://live.bilibili.com"},
            ) as resp:
                if resp.status != 200:
                    logger.warning("[LiveStatus] HTTP %s", resp.status)
                    await asyncio.sleep(3)
                    continue
                try:
                    payload = await resp.json(content_type=None)
                except ContentTypeError:
                    text = (await resp.text())[:200]
                    logger.warning("[LiveStatus] 非 JSON，前 200 字：%s", text)
                    await asyncio.sleep(3)
                    continue

            data = payload.get("data") or {}
            now = _now()

            for room_id in ROOM_IDS:
                uid = ROOM_UIDS.get(room_id)
                info = data.get(str(uid)) if uid is not None else None
                prev = LAST_STATUS.get(room_id, 0)

                if not info or "live_status" not in info:
                    # 沿用上一轮
                    continue

                raw_status = int(info.get("live_status", 0))
                status = 0 if raw_status == 2 else raw_status
                LAST_STATUS[room_id] = status

                if status == 1:
                    live_time_raw = info.get("live_time", 0)
                    try:
                        if isinstance(live_time_raw, (int, float)) or str(live_time_raw).isdigit():
                            start_dt = datetime.datetime.fromtimestamp(int(live_time_raw))
                        else:
                            start_dt = now
                    except (ValueError, OSError, OverflowError):
                        start_dt = now

                    raw_title = info.get("title") or ""
                    LIVE_INFO[room_id]["live_time"] = start_dt.strftime("%Y-%m-%d %H:%M:%S")
                    LIVE_INFO[room_id]["title"] = raw_title

                    if prev == 0:
                        logger.info("[LiveStatus] room_id=%s 上播 %s", room_id, LIVE_INFO[room_id]["live_time"])
                else:
                    LIVE_INFO[room_id]["live_time"] = "0000-00-00 00:00:00"
                    LIVE_INFO[room_id]["title"] = ""
                    if prev == 1:
                        logger.info("[LiveStatus] room_id=%s 下播", room_id)

        except Exception as e:
            logger.error("[LiveStatus] 调用异常: %s", e)

        await asyncio.sleep(3)

@dataclass(slots=True)
class CaptainEvent:
    uid: str
    username: str
    level: str
    ship_count: int
    is_red_packet: bool
    joined_at: datetime.datetime


CAPTAIN_QUEUE: "asyncio.Queue[CaptainEvent]" = asyncio.Queue(maxsize=5000)

def _level_name(guard_level: Any) -> Optional[str]:
    try:
        gl = int(guard_level)
    except Exception:
        return None
    # bilibili: 3=舰长, 2=提督, 1=总督
    if gl == 3:
        return "舰长"
    if gl == 2:
        return "提督"
    if gl == 1:
        return "总督"
    return None


async def _fetch_guard_page(page: int) -> list[dict[str, Any]]:
    if aiohttp_session is None:
        logger.error("[Guard] aiohttp_session 未初始化，无法拉取在舰列表")
        return []

    params = {
        "ruid": GUARD_REPORT_RUID,
        "roomid": GUARD_REPORT_ROOM_ID,
        "page": page,
        "page_size": 30,
        "typ": 5,
    }
    try:
        async with aiohttp_session.get(
            GUARD_LIST_API,
            params=params,
            timeout=10,
            headers={
                "User-Agent": USER_AGENT,
                "Referer": f"https://live.bilibili.com/{GUARD_REPORT_ROOM_ID}",
            },
        ) as resp:
            if resp.status != 200:
                logger.warning("[Guard] page=%s HTTP %s", page, resp.status)
                return []
            try:
                payload = await resp.json(content_type=None)
            except ContentTypeError:
                text = (await resp.text())[:200]
                logger.warning("[Guard] page=%s 非 JSON，前 200 字：%s", page, text)
                return []
    except Exception as e:
        logger.error("[Guard] page=%s 请求异常: %s", page, e)
        return []

    data = payload.get("data") or {}
    raw_list = data.get("list") or []
    if not isinstance(raw_list, list):
        return []
    return raw_list


def _normalize_guard_rows(raw_list: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in raw_list:
        uinfo = item.get("uinfo") or {}
        base = uinfo.get("base") or {}
        guard = uinfo.get("guard") or {}
        uid = uinfo.get("uid") or base.get("uid") or ""
        name = base.get("name") or ""
        level = _level_name(guard.get("level")) or ""
        rows.append({"uid": uid, "name": name, "level": level})
    return rows


async def _collect_guard_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    page = 1
    while True:
        raw_list = await _fetch_guard_page(page)
        if not raw_list:
            break
        rows.extend(_normalize_guard_rows(raw_list))
        page += 1
        await asyncio.sleep(0.8)
    return rows


def _build_guard_report_xlsx(rows: list[dict[str, Any]], month: str) -> Path:
    GUARD_REPORT_ROOT.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{month}在舰列表"

    headers = ["UID", "用户名", "舰长等级"]
    worksheet.append(headers)

    column_widths = [len(header) for header in headers]
    for row in rows:
        values = [row.get("uid", ""), row.get("name", ""), row.get("level", "")]
        worksheet.append(values)
        for index, value in enumerate(values):
            value_length = len(str(value))
            if value_length > column_widths[index]:
                column_widths[index] = value_length

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width + 2

    filename = f"{month}在舰列表.xlsx"
    file_path = GUARD_REPORT_ROOT / filename
    workbook.save(file_path)
    return file_path


def _send_guard_report_email(file_path: Path, month: str, total: int) -> None:
    settings = get_settings()
    if not (settings.smtp_host and settings.email_from):
        logger.warning("[SMTP] 未配置 SMTP_HOST/EMAIL_FROM，跳过在舰列表邮件")
        return

    subject = f"{month}在舰列表（自动发送）"
    body = f"{month} 在舰列表已生成，共 {total} 人。"

    msg = MIMEMultipart()
    msg["Subject"] = subject
    msg["From"] = settings.email_from
    msg["To"] = GUARD_REPORT_RECEIVER
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with file_path.open("rb") as f:
        attachment = MIMEApplication(
            f.read(),
            _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    attachment.add_header("Content-Disposition", "attachment", filename=file_path.name)
    msg.attach(attachment)

    if int(settings.smtp_port) == 465:
        with smtplib.SMTP_SSL(settings.smtp_host, settings.smtp_port) as server:
            if settings.smtp_user and settings.smtp_pass:
                server.login(settings.smtp_user, settings.smtp_pass)
            server.sendmail(settings.email_from, [GUARD_REPORT_RECEIVER], msg.as_string())
    else:
        with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
            server.starttls()
            if settings.smtp_user and settings.smtp_pass:
                server.login(settings.smtp_user, settings.smtp_pass)
            server.sendmail(settings.email_from, [GUARD_REPORT_RECEIVER], msg.as_string())


async def _run_guard_report(target_time: datetime.datetime) -> None:
    month = target_time.strftime("%Y%m")
    logger.info("[Guard] 开始拉取 %s 在舰列表", month)
    rows = await _collect_guard_rows()
    file_path = _build_guard_report_xlsx(rows, month)
    await asyncio.to_thread(_send_guard_report_email, file_path, month, len(rows))
    logger.info("[Guard] 在舰列表已发送 month=%s count=%s path=%s", month, len(rows), file_path)


async def guard_report_scheduler() -> None:
    while True:
        now = _now()
        target = _next_month_end_minute(now)
        sleep_sec = max(1.0, (target - now).total_seconds())
        logger.info("[Guard] 距离下一次在舰列表任务还有约 %.1f 分钟", sleep_sec / 60.0)
        await asyncio.sleep(sleep_sec)

        try:
            await _run_guard_report(target)
        except asyncio.CancelledError:
            logger.debug("[Guard] 在舰列表任务 Cancelled（预期）")
        except Exception as e:
            logger.error("[Guard] 在舰列表任务失败: %s", e)


async def captain_writer_worker() -> None:
    while True:
        ev = await CAPTAIN_QUEUE.get()
        try:
            async with async_session_factory() as session:
                row = Captain(
                    user_uid=ev.uid,
                    username=ev.username,
                    joined_at=ev.joined_at,
                    joined_month=month_str(ev.joined_at),
                    level=ev.level,
                    ship_count=int(ev.ship_count or 1),
                    is_red_packet=bool(ev.is_red_packet),
                )
                session.add(row)
                await session.commit()
        except Exception as e:
            logger.error("[Captain] 写入失败 uid=%s level=%s: %s", ev.uid, ev.level, e)
            try:
                await session.rollback()
            except Exception:
                pass
        finally:
            CAPTAIN_QUEUE.task_done()

async def _start_client(room_id: int) -> None:
    client = blivedm.BLiveClient(room_id, session=aiohttp_session)
    client.set_handler(MyHandler())
    client.start()
    ROOM_CLIENTS[room_id] = client
    LAST_RECONNECT.setdefault(room_id, _now() - datetime.timedelta(days=random.random() * 3.0))
    logger.info("[connect] 已连接房间 %s", room_id)


async def run_clients_loop() -> None:
    for room_id in ROOM_IDS:
        await _start_client(room_id)
        await asyncio.sleep(3)


async def _reconnect_one(room_id: int) -> None:
    if LAST_STATUS.get(room_id, 0) == 1:
        logger.info("[reconnect] room_id=%s 在播，跳过重连", room_id)
        return

    client = ROOM_CLIENTS.get(room_id)
    if client is not None:
        try:
            await client.stop_and_close()
        except asyncio.CancelledError:
            logger.debug("[reconnect] room_id=%s stop_and_close Cancelled（预期）", room_id)
        except Exception as e:
            logger.warning("[reconnect] room_id=%s stop_and_close 异常: %s", room_id, e)

    await asyncio.sleep(3)
    await _start_client(room_id)
    LAST_RECONNECT[room_id] = _now()
    logger.info("[reconnect] room_id=%s 重连完成", room_id)


async def reconnect_scheduler() -> None:
    while True:
        now = _now()
        target = now.replace(hour=6, minute=0, second=0, microsecond=0)
        if now >= target:
            target += datetime.timedelta(days=1)
        sleep_sec = max(1.0, (target - now).total_seconds())
        logger.info("[reconnect] 距离下一次全量重连还有约 %.1f 分钟", sleep_sec / 60.0)
        await asyncio.sleep(sleep_sec)

        logger.info("[reconnect] 开始执行每日全量重连任务")
        for room_id in ROOM_IDS:
            if LAST_STATUS.get(room_id, 0) == 1:
                logger.info("[reconnect] room_id=%s 在播，跳过", room_id)
                continue
            try:
                await _reconnect_one(room_id)
            except asyncio.CancelledError:
                logger.debug("[reconnect] room_id=%s Cancelled（预期）", room_id)
            except Exception as e:
                logger.error("[reconnect] room_id=%s 重连失败: %s", room_id, e)
            await asyncio.sleep(3 + random.uniform(0.5, 2.0))
        logger.info("[reconnect] 本日全量重连任务完成")


class MyHandler(blivedm.BaseHandler):
    def _on_heartbeat(self, client: blivedm.BLiveClient, message: Any) -> None:  # noqa: N802
        return

    def _on_gift(self, client: blivedm.BLiveClient, message: Any) -> None:  # noqa: N802
        try:
            uid = getattr(message, "uid", 0) or 0
            if int(uid) == 0:
                _send_cookie_invalid_email_async(f"[{client.room_id}] gift uid=0")
                return

            gift_name = getattr(message, "gift_name", "") or ""
            if gift_name != "口水黄豆":
                return

            num = getattr(message, "num", 1) or 1
            username = getattr(message, "uname", None) or getattr(message, "username", None) or ""
            try:
                loop = asyncio.get_running_loop()
            except RuntimeError:
                logger.warning("[Gift] 未找到运行中的事件循环，跳过记录 uid=%s", uid)
                return

            loop.create_task(
                _record_gift_ranking(uid=str(uid), username=str(username), gift_count=int(num))
            )
            logger.info(
                "[%s] %s 送了 %s 个口水黄豆",
                client.room_id,
                username,
                num,
            )
        except Exception as e:
            logger.error("[Gift] 处理礼物记录异常: %s", e)

    def _on_user_toast_v2(self, client: blivedm.BLiveClient, message: Any) -> None:  # noqa: N802
        try:
            uid = getattr(message, "uid", 0) or 0
            if int(uid) == 0:
                _send_cookie_invalid_email_async(f"[{client.room_id}] user_toast uid=0")
                return

            username = getattr(message, "username", None) or getattr(message, "uname", None) or ""
            level = _level_name(getattr(message, "guard_level", None))
            if not level:
                return

            num = getattr(message, "num", 1) or 1
            price = getattr(message, "price", 0)
            try:
                is_red_packet = int(price) == 1900
            except Exception:
                is_red_packet = False

            ev = CaptainEvent(
                uid=str(uid),
                username=str(username),
                level=level,
                ship_count=int(num),
                is_red_packet=is_red_packet,
                joined_at=_now(),
            )

            try:
                CAPTAIN_QUEUE.put_nowait(ev)
            except asyncio.QueueFull:
                logger.warning("[Captain] 队列已满，丢弃 uid=%s room_id=%s", ev.uid, client.room_id)
        except Exception as e:
            logger.error("[Captain] 处理 USER_TOAST_V2 异常: %s", e)

async def _record_gift_ranking(uid: str, username: str, gift_count: int) -> None:
    async with async_session_factory() as session:
        try:
            stmt = insert(GiftRanking).values(
                user_uid=uid,
                username=username or None,
                gift_count=gift_count,
            )
            update_values = {
                "gift_count": GiftRanking.gift_count + gift_count,
            }
            if username:
                update_values["username"] = username
            stmt = stmt.on_duplicate_key_update(**update_values)
            await session.execute(stmt)
            await session.commit()
        except Exception as e:
            logger.error("[Gift] 写入 gift_ranking 失败 uid=%s: %s", uid, e)
            try:
                await session.rollback()
            except Exception:
                pass


def live_status_snapshot() -> Dict[str, Any]:
    room_id = ROOM_IDS[0] if ROOM_IDS else 0
    info = LIVE_INFO.get(room_id, {}) if room_id else {}
    return {
        "status": LAST_STATUS.get(room_id, 0) if room_id else 0,
        "live_time": info.get("live_time", "0000-00-00 00:00:00"),
        "title": info.get("title", ""),
    }

_tasks: list[asyncio.Task] = []


async def bootstrap() -> None:
    settings = get_settings()
    if not settings.bili_monitor_enabled:
        logger.info("[bili] BILI_MONITOR_ENABLED 未开启，跳过监听启动")
        return

    init_session()

    try:
        await init_uids_once()
    except Exception as e:
        logger.error("[bili] init_uids_once 失败: %s", e)

    loop = asyncio.get_running_loop()
    _tasks.extend(
        [
            loop.create_task(captain_writer_worker(), name="bili:cpt_writer"),
            loop.create_task(run_clients_loop(), name="bili:clients"),
            loop.create_task(monitor_all_rooms_status(), name="bili:live_status"),
            loop.create_task(reconnect_scheduler(), name="bili:reconnect"),
            loop.create_task(guard_report_scheduler(), name="bili:guard_report"),
        ]
    )
    logger.info("[bili] 后台监听已启动，tasks=%d", len(_tasks))


async def shutdown() -> None:
    tasks = list(_tasks)
    for t in tasks:
        t.cancel()
    _tasks.clear()
    if tasks:
        await asyncio.gather(*tasks, return_exceptions=True)

    # 关闭客户端
    for rid, client in list(ROOM_CLIENTS.items()):
        try:
            await client.stop_and_close()
        except Exception:
            pass
    ROOM_CLIENTS.clear()

    global aiohttp_session
    if aiohttp_session:
        try:
            await aiohttp_session.close()
        except Exception:
            pass
        aiohttp_session = None
