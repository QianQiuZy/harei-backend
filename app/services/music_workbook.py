from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Final, Protocol, TypeAlias
from urllib.parse import urlsplit
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from app.models.music import Song


IMPORT_SHEET: Final = "导入数据"
SONG_LIST_SHEET: Final = "歌曲列表"
IMPORT_HEADERS: Final = ("歌名", "日期", "直播标题", "歌切链接")
ExcelValue: TypeAlias = (
    bool
    | int
    | float
    | Decimal
    | str
    | CellRichText
    | datetime
    | date
    | time
    | timedelta
    | DataTableFormula
    | ArrayFormula
    | None
)


class WorkbookSheetFactory(Protocol):
    def create_sheet(self, title: str, index: int | None = None) -> Worksheet: ...


@dataclass(frozen=True, slots=True)
class PerformanceImportRow:
    excel_row: int
    song_title: str
    performed_on: date
    stream_title: str
    clip_url: str


@dataclass(frozen=True, slots=True)
class WorkbookIssue:
    row: int
    field: str
    code: str
    message: str

    def as_dict(self) -> dict[str, int | str]:
        return {
            "row": self.row,
            "field": self.field,
            "code": self.code,
            "message": self.message,
        }


def build_performance_template(songs: list[Song]) -> BytesIO:
    workbook = Workbook()
    import_sheet = workbook.active
    if import_sheet is None:
        raise RuntimeError("Workbook has no active worksheet")
    import_sheet.title = IMPORT_SHEET
    import_sheet.append(IMPORT_HEADERS)
    import_sheet.freeze_panes = "A2"
    import_sheet.auto_filter.ref = "A1:D1"
    _style_header(import_sheet)
    _set_widths(import_sheet, (28, 16, 32, 56))

    song_sheet = _create_sheet(workbook, SONG_LIST_SHEET)
    song_sheet.append(("歌名", "歌手", "状态"))
    for song in songs:
        song_sheet.append((song.title, song.artist, "公开" if song.status == "active" else "归档"))
    song_sheet.freeze_panes = "A2"
    song_sheet.auto_filter.ref = f"A1:C{max(song_sheet.max_row, 1)}"
    _style_header(song_sheet)
    _set_widths(song_sheet, (32, 32, 12))

    output = BytesIO()
    _ = workbook.save(output)
    _ = output.seek(0)
    return output


def parse_performance_workbook(contents: bytes) -> tuple[list[PerformanceImportRow], list[WorkbookIssue]]:
    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
    except (BadZipFile, OSError, ValueError, KeyError) as exc:
        return [], [WorkbookIssue(1, "file", "INVALID_XLSX", f"无法读取 XLSX：{exc}")]

    if IMPORT_SHEET not in workbook.sheetnames:
        return [], [WorkbookIssue(1, "sheet", "MISSING_SHEET", f"缺少工作表“{IMPORT_SHEET}”")]
    sheet = workbook[IMPORT_SHEET]
    values = sheet.iter_rows(values_only=True)
    header = next(values, None)
    if tuple(header or ()) != IMPORT_HEADERS:
        return [], [WorkbookIssue(1, "header", "INVALID_HEADERS", "表头必须依次为：歌名、日期、直播标题、歌切链接")]

    rows: list[PerformanceImportRow] = []
    issues: list[WorkbookIssue] = []
    seen: set[tuple[str, date, str]] = set()
    for excel_row, raw in enumerate(values, start=2):
        if not raw or all(value is None or value == "" for value in raw):
            continue
        title = _required_text(raw[0] if len(raw) > 0 else None)
        performed_on = _date_value(raw[1] if len(raw) > 1 else None)
        stream_title = _required_text(raw[2] if len(raw) > 2 else None)
        clip_url = _required_text(raw[3] if len(raw) > 3 else None)
        row_issues = _row_issues(excel_row, title, performed_on, stream_title, clip_url)
        if row_issues:
            issues.extend(row_issues)
            continue
        assert title is not None and performed_on is not None and stream_title is not None and clip_url is not None
        key = (title, performed_on, clip_url)
        if key in seen:
            issues.append(WorkbookIssue(excel_row, "row", "DUPLICATE_ROW", "文件内存在重复演唱记录"))
            continue
        seen.add(key)
        rows.append(PerformanceImportRow(excel_row, title, performed_on, stream_title, clip_url))
    if not rows and not issues:
        issues.append(WorkbookIssue(2, "row", "EMPTY_IMPORT", "导入数据中没有可导入记录"))
    return rows, issues


def _required_text(value: ExcelValue) -> str | None:
    match value:
        case None | "" | ArrayFormula() | DataTableFormula():
            return None
        case str() | CellRichText() | bool() | int() | float() | Decimal() | datetime() | date() | time() | timedelta():
            return str(value).strip() or None


def _date_value(value: ExcelValue) -> date | None:
    match value:
        case datetime():
            return value.date()
        case date():
            return value
        case str():
            try:
                return date.fromisoformat(value.strip())
            except ValueError:
                return None
        case None | bool() | int() | float() | Decimal() | CellRichText() | time() | timedelta() | DataTableFormula() | ArrayFormula():
            return None


def _row_issues(
    row: int,
    title: str | None,
    performed_on: date | None,
    stream_title: str | None,
    clip_url: str | None,
) -> list[WorkbookIssue]:
    issues: list[WorkbookIssue] = []
    if title is None:
        issues.append(WorkbookIssue(row, "歌名", "REQUIRED", "歌名不能为空"))
    if performed_on is None:
        issues.append(WorkbookIssue(row, "日期", "INVALID_DATE", "日期必须是 Excel 日期或 YYYY-MM-DD"))
    if stream_title is None:
        issues.append(WorkbookIssue(row, "直播标题", "REQUIRED", "直播标题不能为空"))
    if clip_url is None:
        issues.append(WorkbookIssue(row, "歌切链接", "REQUIRED", "歌切链接不能为空"))
    elif urlsplit(clip_url).scheme not in {"http", "https"} or not urlsplit(clip_url).netloc:
        issues.append(WorkbookIssue(row, "歌切链接", "INVALID_URL", "歌切链接必须使用 http 或 https"))
    return issues


def _style_header(sheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="282837")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")


def _set_widths(sheet: Worksheet, widths: tuple[int, ...]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def _create_sheet(workbook: WorkbookSheetFactory, title: str) -> Worksheet:
    return workbook.create_sheet(title)
