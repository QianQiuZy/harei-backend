from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from redis.asyncio import Redis
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.redis import get_redis_client
from app.db.session import get_db_session
from app.deps.auth import get_bearer_token
from app.models.captain import Captain
from app.schemas.captains import CaptainItem, CaptainListResponse
from app.services.auth_service import AuthService

router = APIRouter()


async def require_token(
    token: str = Depends(get_bearer_token),
    redis: Redis = Depends(get_redis_client),
) -> None:
    service = AuthService(redis)
    username = await service.get_username_by_token(token)
    if not username:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")


@router.get("/captains", response_model=CaptainListResponse)
async def list_captains(
    month: str | None = Query(default=None),
    uid: str | None = Query(default=None),
    session: AsyncSession = Depends(get_db_session),
    _: None = Depends(require_token),
) -> CaptainListResponse:
    current_month = datetime.utcnow().strftime("%Y%m")
    target_month = month or current_month

    query = select(Captain)
    if uid:
        query = query.where(Captain.user_uid == uid)
    else:
        query = query.where(Captain.joined_month == target_month)

    result = await session.execute(query.order_by(Captain.joined_at.asc()))
    rows = result.scalars().all()

    items = [
        CaptainItem(
            uid=row.user_uid,
            name=row.username,
            level=row.level,
            count=row.ship_count,
            red_packet=row.is_red_packet,
            joined_at=row.joined_at,
        )
        for row in rows
    ]
    return CaptainListResponse(items=items)


@router.get("/captains/xlsx")
async def export_captains_xlsx(
    month: str = Query(..., pattern=r"^\d{6}$"),
    session: AsyncSession = Depends(get_db_session),
    _: None = Depends(require_token),
) -> StreamingResponse:
    query = select(Captain).where(Captain.joined_month == month)
    result = await session.execute(query.order_by(Captain.joined_at.asc()))
    rows = result.scalars().all()

    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"未找到{month}上舰记录")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{month}舰长"

    headers = ["UID", "用户名", "舰长等级", "上舰数量", "上舰时间", "是否红包"]
    worksheet.append(headers)

    column_widths = [len(header) for header in headers]
    for row in rows:
        joined_at = row.joined_at.strftime("%Y-%m-%d %H:%M:%S")
        values = [
            row.user_uid,
            row.username or "",
            row.level,
            row.ship_count,
            joined_at,
            "是" if row.is_red_packet else "否",
        ]
        worksheet.append(values)
        for index, value in enumerate(values):
            value_length = len(str(value))
            if value_length > column_widths[index]:
                column_widths[index] = value_length

    for index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width + 2

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"captains_{month}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
